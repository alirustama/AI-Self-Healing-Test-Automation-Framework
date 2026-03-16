"""
Excel sheet reader utility for test data management
Handles reading contact data from Excel files
"""

from pathlib import Path
import openpyxl
from typing import List, Dict, Any
from utils.logger import get_logger

logger = get_logger(__name__)


class SheetReader:
    """Reads and parses Excel files for test data"""
    
    def __init__(self, file_path: str):
        """
        Initialize SheetReader with an Excel file
        
        Args:
            file_path (str): Path to the Excel file
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            logger.error(f"Excel file not found: {file_path}")
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        logger.info(f"Loading Excel file: {file_path}")
        self.workbook = openpyxl.load_workbook(self.file_path)
    
    def get_sheet_names(self) -> List[str]:
        """Get all sheet names in the workbook"""
        return self.workbook.sheetnames
    
    def read_sheet(self, sheet_name: str = None) -> List[Dict[str, Any]]:
        """
        Read data from a sheet and return as list of dictionaries
        
        Args:
            sheet_name (str): Name of the sheet to read. If None, uses first sheet.
        
        Returns:
            List[Dict]: List of row dictionaries
        """
        if sheet_name is None:
            sheet_name = self.workbook.sheetnames[0]
        
        logger.info(f"Reading sheet: {sheet_name}")
        sheet = self.workbook[sheet_name]
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # Read data rows
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {}
            for col_idx, header in enumerate(headers):
                if header:
                    row_dict[header] = row[col_idx] if col_idx < len(row) else None
            
            if any(row_dict.values()):  # Skip empty rows
                data.append(row_dict)
                logger.debug(f"Row {row_idx}: {row_dict}")
        
        logger.info(f"Successfully read {len(data)} rows from {sheet_name}")
        return data
    
    def read_contacts(self, sheet_name: str = "Contacts") -> List[Dict[str, Any]]:
        """
        Read contact data from Excel
        
        Args:
            sheet_name (str): Name of the contacts sheet
        
        Returns:
            List[Dict]: List of contact dictionaries
        """
        try:
            contacts = self.read_sheet(sheet_name)
            logger.info(f"Loaded {len(contacts)} contacts")
            return contacts
        except KeyError as e:
            logger.error(f"Sheet '{sheet_name}' not found. Available sheets: {self.get_sheet_names()}")
            raise
    
    def close(self):
        """Close the workbook"""
        self.workbook.close()
        logger.info("Workbook closed")


def load_contacts(file_path: str) -> List[Dict[str, Any]]:
    """
    Convenience function to load contacts from Excel
    
    Args:
        file_path (str): Path to the Excel file
    
    Returns:
        List[Dict]: List of contact dictionaries
    """
    reader = SheetReader(file_path)
    contacts = reader.read_contacts()
    reader.close()
    return contacts
